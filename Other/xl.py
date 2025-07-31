import re
from copy import copy
from openpyxl import load_workbook

wb = load_workbook(r"C:\Users\OOPPEENN\Desktop\2dfan.xlsx")
url_pattern = re.compile(r'https?://')

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and url_pattern.match(cell.value):
                # 备份所有需保留的属性
                orig_fill      = copy(cell.fill)
                orig_border    = copy(cell.border)
                orig_alignment = copy(cell.alignment)
                orig_number    = cell.number_format
                orig_protect   = copy(cell.protection)
                orig_font      = copy(cell.font)
                orig_comment   = copy(cell.comment)

                # 设置超链接（这一步往往会重置样式，所以先备份）
                cell.hyperlink = cell.value
                cell.value     = cell.value

                # 还原样式
                cell.fill          = orig_fill
                cell.border        = orig_border
                cell.alignment     = orig_alignment
                cell.number_format = orig_number
                cell.protection    = orig_protect
                cell.font          = orig_font
                cell.comment       = orig_comment

wb.save(r"C:\Users\OOPPEENN\Desktop\2dfan_o.xlsx")